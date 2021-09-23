import random
import openpyxl as xls


def roll_student():

    path = 'C:/Users/pr0mo/Desktop/Klasy.xlsx'
    wb = xls.load_workbook(path)
    sheets = wb.sheetnames
    uniq = []
    try:
        to_ask = int(input("Hey input how many students do you want to ask! \n"))
        Class = int(input("Hey choose class: 1 - I C , 2 - II C , 3 - III C \n"))
    except ValueError:
        print("That's not an int!")
    else:

        if Class == 1:
            ws = wb[sheets[0]]
        elif Class == 2:
            ws = wb[sheets[1]]
        elif Class == 3:
            ws = wb[sheets[2]]
        else:
            print("Brr wrong Class")
            exit()

        all_students = ws.max_row
        ask_students = 0
        if to_ask > all_students:
            print(f"Error Class only count {all_students} people!")
            return
        else:
            pass
        while ask_students != to_ask:
            pick = random.randint(2, all_students)
            if pick not in uniq:
                uniq.append(pick)
                name = ws[f'B{pick}'].value
                surname = ws[f'C{pick}'].value
                ask_students += 1
                students_to_ask = [name] + [surname]
                print(students_to_ask)
